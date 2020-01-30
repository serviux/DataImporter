import openpyxl
import player
import datetime
import json
import os.path
class ExcelInterface:
    """
    This class is used to interface with the excel document cointaining all the players
    matches and win loss ratios.

    args: path_to_workbook --path to the file being taken in.
    """
    def __init__(self, path_to_workbook):
        if len(path_to_workbook) == 0:
            raise Exception("No workbook loaded")
        self.wb = openpyxl.load_workbook(path_to_workbook)
    
    #stores the sheets in the current instance
    def load_sheets(self):
        self.sheets = self.wb.sheetnames

    """
    gets a list of all players from all tabs in the excel doc
    and adds them to a dictionary. Names are forced to be uppercase 
    so they are case Insensitive in the word doc

    creates the self.entries dictionary a dictionary 
    containing a key value pair like <String, Player>
    """
    def get_entries(self): 
        
        self.entries = dict()

        for name in self.sheets:
            ws = self.wb[name]
            entries_weekly = ws["Q"]

            skipHeader = True
            for val in entries_weekly:
                # ignore the 'Entries' 
                if skipHeader :
                    skipHeader = False
                # only add into dictionary if name is
                elif val.value is None:
                    break
                elif str.upper(val.value) not in self.entries:
                    self.entries[str.upper(val.value)] =  player.Player(name = str.upper(val.value))
    
    """
    private function to increment player data
    increments total wins/losses
    adds players to the last 5 plays list
    """

    def _append_player_data(self, names, match1, match2):
        
        #loop through each pair of players, step through 2 at a time
        # players in matches are adjacent to each other in the list
        for j in range(0, len(names) , 2):
            
            if names[j] is None or match1[j] is None or match2[j] is None:
                break
            name1 = str.upper(names[j])
            name2 = str.upper(names[j+1])
            #Player one in the match
            if match1[j] is not None:
                self.entries[name1].wins += 1
            else:
                self.entries[name1].losses += 1

            
            if match2[j] is not None:
                self.entries[name1].wins += 1
            else:
                self.entries[name1].losses += 1
            
            #Player 2 in the match
            k = j + 1
            if match1[k] is not None:
                self.entries[name2].wins += 1
            else:
                self.entries[name2].losses += 1     

            if match2[k] is not None:
                self.entries[name2].wins += 1
            else:
                self.entries[name2].losses += 1

            #add both players to last played list
            self.entries[name1].add_last_played(name2)
            self.entries[name2].add_last_played(name1)
    """
    find the scores of the matches that each player fought through the semester
    """
    def find_scores(self):
        
        for name in self.sheets:
            ws = self.wb[name]
            max_row = ws.max_row
            max_col = ws.max_column
            columns = ws.iter_cols(min_col= 1 , min_row= 1 ,
             max_col= max_col, max_row= max_row-4, values_only= False)
             
            i = 1
            names = []
            match1 = []
            match2 = []
            for col in columns:
               
                skipHeader = True
                if i%4 == 1: 
                    for cell in col:
                        if skipHeader:              
                            skipHeader =False
                        else:  
                            names.append(cell.value)
                if i%4 == 2:
                    for cell in col:
                        if skipHeader:
                            skipHeader = False
                        else:            
                            match1.append(cell.value)
                if i%4 == 3:
                    for cell in col:
                        if skipHeader:
                            skipHeader = False
                        else:    
                            match2.append(cell.value)
                if i%4 == 0:
                     
                     #TODO: Store rounds in self.entries
                     self._append_player_data(names, match1, match2)
                     names = []
                     match1 = []
                     match2 = []
                     
                     
                i += 1        
                


                       


        

    """
    used to create a list of all the players converted to a 
    json object. 
    """
    def playersToJSON(self):
        players = []
        for key in self.entries.keys():
            players.append(player.PlayerToJSON.PlayerToJSON(self.entries[key]))
        return players
    
    def store_to_json(self, path):
        data = { "event" : "smash-league",
          "date" : datetime.datetime.now().strftime("%Y/%m/%d"),
          "players" : []
           }
        ps = []
        for player in self.playersToJSON():
            ps.append(json.loads(player))

       
        data["players"] = ps
        string_date = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        file_path = os.path.join(path, "league_data_" + string_date + ".json")
        with open(file_path, "w" ) as fil:
            json.dump(data, fil, indent=4)


if __name__ == "__main__":
    xl = ExcelInterface("Brackets.xlsx")
    xl.load_sheets()
    xl.get_entries()
    xl.find_scores()
    print(xl.playersToJSON())
    xl.store_to_json()



        




